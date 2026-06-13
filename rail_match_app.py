"""
레일연마차 업무자동화
실행: streamlit run rail_match_app.py
"""

import datetime
import io
import re

import openpyxl
import pandas as pd
import streamlit as st

# ── 페이지 설정 ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="레일연마차 업무자동화",
    page_icon="🚃",
    layout="wide",
)

# ── 로그인 ────────────────────────────────────────────────────────────────────

def check_password():
    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <div style="max-width:360px; margin:80px auto 0; text-align:center;">
        <div style="font-size:2.8rem; margin-bottom:8px;">🚃</div>
        <div style="font-size:1.2rem; font-weight:700; color:#1a2332; margin-bottom:4px;">
            레일연마차 업무자동화
        </div>
        <div style="font-size:0.82rem; color:#64748b; margin-bottom:28px;">
            대저궤도장비분소 전용 시스템
        </div>
    </div>
    """, unsafe_allow_html=True)

    col = st.columns([1, 2, 1])[1]
    with col:
        pw = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
        if st.button("로그인", use_container_width=True):
            if pw == st.secrets["PASSWORD"]:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    return False

if not check_password():
    st.stop()

st.markdown("""
<style>
    /* 전체 */
    [data-testid="stAppViewContainer"] { background: #f0f4f8; }

    /* 사이드바 */
    [data-testid="stSidebar"] { background: #1a2332; padding-top: 0; }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] .stFileUploader label {
        color: #94a3b8 !important;
        font-size: 0.75rem !important;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }

    /* 사이드바 앱 타이틀 */
    .sb-brand {
        padding: 20px 16px 14px;
        border-bottom: 1px solid #2d3f55;
        margin-bottom: 8px;
    }
    .sb-brand .sb-icon { font-size: 1.5rem; }
    .sb-brand .sb-title {
        font-size: 0.95rem;
        font-weight: 700;
        color: #f1f5f9 !important;
        margin-top: 4px;
        line-height: 1.3;
    }
    .sb-brand .sb-sub {
        font-size: 0.7rem;
        color: #64748b !important;
        margin-top: 2px;
    }

    /* 사이드바 메뉴 */
    .sb-menu-label {
        font-size: 0.65rem !important;
        color: #475569 !important;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        padding: 14px 16px 4px;
        display: block;
    }
    .sb-menu-item {
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 9px 16px;
        margin: 1px 8px;
        border-radius: 8px;
        cursor: pointer;
        font-size: 0.85rem;
        color: #94a3b8 !important;
        text-decoration: none;
        transition: background 0.15s;
    }
    .sb-menu-item:hover { background: #243347; color: #e2e8f0 !important; }
    .sb-menu-item.active {
        background: #0f766e22;
        color: #5eead4 !important;
        border: 1px solid #0f766e44;
    }
    .sb-menu-item .icon { font-size: 1rem; width: 20px; text-align: center; }

    /* 메인 콘텐츠 */
    .page-header {
        background: linear-gradient(135deg, #1a2332 0%, #2d3f55 100%);
        border-radius: 12px;
        padding: 24px 28px 20px;
        margin-bottom: 20px;
    }
    .page-header .ph-badge {
        display: inline-block;
        background: #0f766e33;
        color: #5eead4 !important;
        font-size: 0.65rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        padding: 3px 10px;
        border-radius: 99px;
        border: 1px solid #0f766e66;
        margin-bottom: 8px;
    }
    .page-header h2 {
        margin: 0;
        font-size: 1.4rem;
        font-weight: 700;
        color: white !important;
        letter-spacing: -0.02em;
    }
    .page-header p {
        margin: 6px 0 0;
        font-size: 0.82rem;
        color: #94a3b8 !important;
    }

    /* 업로드 카드 */
    .upload-section {
        background: white;
        border-radius: 12px;
        padding: 20px 24px;
        border: 1px solid #e2e8f0;
        margin-bottom: 16px;
    }
    .upload-section .us-title {
        font-size: 0.7rem;
        font-weight: 700;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 14px;
        padding-bottom: 8px;
        border-bottom: 1px solid #f1f5f9;
    }

    /* 상태 배지 */
    .status-row {
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 8px 12px;
        border-radius: 8px;
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        margin-bottom: 8px;
        font-size: 0.82rem;
    }
    .status-row.ok { background: #f0fdf4; border-color: #bbf7d0; }
    .status-row.wait { background: #fafafa; border-color: #e5e7eb; }
    .dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
    .dot.green { background: #22c55e; }
    .dot.gray { background: #d1d5db; }

    /* 구간 탭 결과 */
    .seg-header {
        display: flex;
        align-items: center;
        gap: 10px;
        margin: 20px 0 10px;
    }
    .seg-badge {
        background: #1a2332;
        color: #e2e8f0 !important;
        font-size: 0.72rem;
        font-weight: 700;
        padding: 4px 12px;
        border-radius: 6px;
        letter-spacing: 0.05em;
    }
    .seg-count {
        font-size: 0.78rem;
        color: #64748b;
    }
    .seg-count.has-match { color: #0f766e; font-weight: 600; }

    /* 지표 카드 */
    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 16px 20px;
        border: 1px solid #e2e8f0;
    }
    .metric-card .mc-label {
        font-size: 0.68rem;
        font-weight: 700;
        color: #94a3b8;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 4px;
    }
    .metric-card .mc-value {
        font-size: 1.9rem;
        font-weight: 800;
        color: #1a2332;
        line-height: 1;
    }
    .metric-card .mc-value.green { color: #0f766e; }
    .metric-card .mc-value.amber { color: #b45309; }
    .metric-card .mc-value.sm { font-size: 1.2rem; margin-top: 4px; }

    /* 경고/정보 */
    .info-box {
        background: #eff6ff;
        border-left: 4px solid #3b82f6;
        border-radius: 0 8px 8px 0;
        padding: 12px 16px;
        font-size: 0.83rem;
        color: #1e40af;
        margin: 8px 0;
    }
    .warn-box {
        background: #fffbeb;
        border-left: 4px solid #f59e0b;
        border-radius: 0 8px 8px 0;
        padding: 12px 16px;
        font-size: 0.83rem;
        color: #92400e;
        margin: 8px 0;
    }

    /* 메인 버튼 */
    .stButton > button {
        background: #0f766e;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 600;
        font-size: 0.88rem;
        width: 100%;
    }
    .stButton > button:hover { background: #0d5e58; }

    /* 테이블 */
    [data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

    /* 홈 화면 기능 카드 */
    .feature-card {
        background: white;
        border-radius: 12px;
        padding: 22px 24px;
        border: 1px solid #e2e8f0;
        cursor: pointer;
        transition: box-shadow 0.2s, border-color 0.2s;
        height: 100%;
    }
    .feature-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.08); border-color: #0f766e66; }
    .feature-card .fc-icon { font-size: 1.8rem; margin-bottom: 10px; }
    .feature-card .fc-title { font-size: 1rem; font-weight: 700; color: #1a2332; margin-bottom: 4px; }
    .feature-card .fc-desc { font-size: 0.8rem; color: #64748b; line-height: 1.5; }
    .feature-card .fc-tag {
        display: inline-block;
        margin-top: 10px;
        font-size: 0.65rem;
        font-weight: 700;
        padding: 3px 8px;
        border-radius: 4px;
        background: #f0fdf4;
        color: #0f766e;
        border: 1px solid #bbf7d0;
    }
    .feature-card .fc-tag.soon {
        background: #f8fafc;
        color: #94a3b8;
        border-color: #e2e8f0;
    }
    div[data-testid="stHorizontalBlock"] { align-items: stretch; }
</style>
""", unsafe_allow_html=True)


# ── 핵심 로직 ─────────────────────────────────────────────────────────────────

def parse_pos(val):
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r'^(\d+)[Kk](\d+)$', s)
    if m:
        return int(m.group(1)) * 1000 + int(m.group(2))
    if re.match(r'^\d+(\.\d+)?$', s):
        return int(float(s))
    return None


def overlaps(a0, a1, b0, b1):
    if any(v is None for v in (a0, a1, b0, b1)):
        return False
    return max(min(a0, a1), min(b0, b1)) <= min(max(a0, a1), max(b0, b1))


def load_perf(ws):
    car1 = str(ws['C4'].value or '1호 레일연마차')
    car2 = str(ws['K4'].value or '2호 레일연마차')
    records = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        for car_name, di, seg_i, fi, ti in [(car1, 2, 4, 5, 6), (car2, 10, 12, 13, 14)]:
            d = row[di]
            if d is None or not isinstance(d, (int, float)):
                continue
            fr, to, seg = row[fi], row[ti], row[seg_i]
            records.append({
                'car': car_name,
                'day': int(d),
                'segment': str(seg) if seg is not None else '-',
                'from_raw': fr,
                'to_raw': to,
                'from_m': parse_pos(fr),
                'to_m': parse_pos(to),
            })
    return car1, car2, records


def run_match(ws1, perf_records, year, month):
    results, unmatched = [], []
    current_date = None
    last_start = last_end = last_b = last_c = last_d = None

    for row in ws1.iter_rows(min_row=7, values_only=True):
        a, b, c, d, e, f = row[0], row[1], row[2], row[3], row[4], row[5]
        i_val = row[8]

        if isinstance(a, datetime.datetime):
            current_date = a
        elif isinstance(a, str):
            m = re.match(r'(\d+)/(\d+)', a)
            if m:
                try:
                    current_date = datetime.datetime(year, int(m.group(1)), int(m.group(2)))
                except ValueError:
                    pass

        if current_date is None:
            continue

        if b:
            last_b, last_c, last_d = b, c, d
            last_start, last_end = e, f
        elif e:
            last_start = e
            last_end = f if f else last_end

        if not i_val or '레일연마' not in str(i_val):
            continue

        use_start = e if e is not None else last_start
        use_end   = f if f is not None else last_end
        start_m   = parse_pos(use_start)
        end_m     = parse_pos(use_end)
        day       = current_date.day

        matched = False
        for p in perf_records:
            if p['day'] != day:
                continue
            if overlaps(start_m, end_m, p['from_m'], p['to_m']):
                matched = True
                results.append({
                    '일자': current_date.strftime('%Y. %m. %d.'),
                    '연마차명': p['car'],
                    '공종': str(i_val),
                    '분소': str(last_b or '-'),
                    '일정_시점': str(use_start) if use_start is not None else '-',
                    '일정_종점': str(use_end)   if use_end   is not None else '-',
                    '실적_구간': p['segment'],
                    '실적_부터': str(p['from_raw']),
                    '실적_까지': str(p['to_raw']),
                })

        if not matched and (start_m or end_m):
            unmatched.append({
                '일자': current_date.strftime('%Y. %m. %d.'),
                '공종': str(i_val),
                '분소': str(last_b or '-'),
                '일정_시점': str(use_start) if use_start is not None else '-',
                '일정_종점': str(use_end)   if use_end   is not None else '-',
                '사유': '해당 날짜 실적에 겹치는 구간 없음',
            })

    return results, unmatched


def extract_year_month(filename: str):
    m = re.search(r'(\d{4})년.?(\d{1,2})월', filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    now = datetime.datetime.now()
    return now.year, now.month


COLUMN_CONFIG = {
    '일자':      st.column_config.TextColumn('일자', width=130),
    '연마차명':  st.column_config.TextColumn('연마차명', width=150),
    '공종':      st.column_config.TextColumn('공종', width=120),
    '분소':      st.column_config.TextColumn('분소', width=80),
    '일정_시점': st.column_config.TextColumn('일정 시점', width=100),
    '일정_종점': st.column_config.TextColumn('일정 종점', width=100),
    '실적_구간': st.column_config.TextColumn('실적 구간', width=140),
    '실적_부터': st.column_config.TextColumn('실적 부터', width=100),
    '실적_까지': st.column_config.TextColumn('실적 까지', width=100),
}

# ── 세션 초기화 ───────────────────────────────────────────────────────────────

if 'page' not in st.session_state:
    st.session_state['page'] = 'home'


# ── 사이드바 ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div class="sb-brand">
        <div class="sb-icon">🚃</div>
        <div class="sb-title">레일연마차<br>업무자동화</div>
        <div class="sb-sub">부산도시철도 궤도관리</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<span class="sb-menu-label">메뉴</span>', unsafe_allow_html=True)

    pages = [
        ('home',  '🏠', '홈'),
        ('match', '🔍', '일정-실적 매칭'),
    ]
    for key, icon, label in pages:
        is_active = st.session_state['page'] == key
        cls = 'sb-menu-item active' if is_active else 'sb-menu-item'
        if st.button(f"{icon}  {label}", key=f"nav_{key}",
                     use_container_width=True,
                     type="primary" if is_active else "secondary"):
            st.session_state['page'] = key
            st.rerun()


# ── 홈 화면 ───────────────────────────────────────────────────────────────────

def page_home():
    st.markdown("""
    <div class="page-header">
        <div class="ph-badge">레일연마차 업무자동화</div>
        <h2>🚃 업무자동화 시스템</h2>
        <p>레일연마차 운영 관련 반복 업무를 자동화합니다. 아래에서 기능을 선택하세요.</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="fc-icon">🔍</div>
            <div class="fc-title">일정-실적 매칭</div>
            <div class="fc-desc">궤도정비공사 일정과 레일연마차 작업실적을 비교해 구간 겹침을 자동 확인합니다.</div>
            <span class="fc-tag">사용 가능</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("바로가기 →", key="go_match"):
            st.session_state['page'] = 'match'
            st.rerun()

    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="fc-icon">📊</div>
            <div class="fc-title">월간 실적 보고서</div>
            <div class="fc-desc">레일연마차 월간 작업실적을 자동 집계하고 보고서 형태로 출력합니다.</div>
            <span class="fc-tag soon">준비 중</span>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="fc-icon">🗓️</div>
            <div class="fc-title">작업 일정 관리</div>
            <div class="fc-desc">연마차별 작업 일정을 캘린더 형태로 시각화하고 이동일·검수일을 관리합니다.</div>
            <span class="fc-tag soon">준비 중</span>
        </div>
        """, unsafe_allow_html=True)


# ── 매칭 화면 ─────────────────────────────────────────────────────────────────

def page_match():
    st.markdown("""
    <div class="page-header">
        <div class="ph-badge">일정-실적 매칭</div>
        <h2>🔍 일정-실적 구간 매칭</h2>
        <p>두 파일을 업로드하면 1구간·2구간을 자동으로 매칭합니다.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── 파일 업로드 ───────────────────────────────────────────────────────────
    col_up1, col_up2 = st.columns(2)

    with col_up1:
        st.markdown('<div class="upload-section"><div class="us-title">📋 파일 업로드</div>', unsafe_allow_html=True)
        f_schedule = st.file_uploader(
            "궤도정비공사일정 파일",
            type=['xlsx'], key="schedule",
            help="0000년 0월 궤도시설 보수정비공사 일정.xlsx",
        )
        f_perf = st.file_uploader(
            "레일연마차 작업 및 검수실적 파일",
            type=['xlsx'], key="perf",
            help="1. 0000년 0월 레일연마차 작업 및 검수실적.xlsx",
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with col_up2:
        st.markdown('<div class="upload-section"><div class="us-title">📂 업로드 상태</div>', unsafe_allow_html=True)

        if f_schedule:
            st.markdown(f'<div class="status-row ok"><div class="dot green"></div><b>일정 파일</b> — {f_schedule.name}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="status-row wait"><div class="dot gray"></div>궤도정비공사일정 파일 대기 중</div>', unsafe_allow_html=True)

        if f_perf:
            st.markdown(f'<div class="status-row ok"><div class="dot green"></div><b>실적 파일</b> — {f_perf.name}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="status-row wait"><div class="dot gray"></div>레일연마차 작업 및 검수실적 파일 대기 중</div>', unsafe_allow_html=True)

        if f_schedule and f_perf:
            year, month = extract_year_month(f_schedule.name)
            st.markdown(f"""
            <div style="margin-top:12px; padding: 10px 14px; background:#f0fdf4;
                        border-radius:8px; border:1px solid #bbf7d0;
                        font-size:0.82rem; color:#0f766e;">
                ✅ 두 파일이 준비됐습니다. <b>{year}년 {month}월</b> 데이터를 분석합니다.<br>
                <span style="font-size:0.75rem; color:#16a34a; margin-top:4px; display:block;">
                1구간 · 2구간 자동 매칭 실행 중...
                </span>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="margin-top:12px; padding:10px 14px; background:#f8fafc;
                        border-radius:8px; border:1px solid #e2e8f0;
                        font-size:0.82rem; color:#94a3b8;">
                두 파일을 모두 업로드하면 자동으로 매칭이 시작됩니다.
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── 두 파일 없으면 중단 ───────────────────────────────────────────────────
    if not f_schedule or not f_perf:
        return

    # ── 매칭 실행 ─────────────────────────────────────────────────────────────
    cache_key = f"{f_schedule.name}_{f_perf.name}"
    if st.session_state.get('match_cache_key') != cache_key:
        try:
            wb1 = openpyxl.load_workbook(io.BytesIO(f_schedule.read()), data_only=True)
            wb2 = openpyxl.load_workbook(io.BytesIO(f_perf.read()),     data_only=True)
        except Exception as e:
            st.error(f"파일을 열 수 없습니다: {e}")
            return

        year, month = extract_year_month(f_schedule.name)
        ws2 = wb2['장비별 세부작업실적']
        _, _, perf_records = load_perf(ws2)

        seg_results = {}
        for sheet in ['1구간', '2구간']:
            if sheet in wb1.sheetnames:
                ws1 = wb1[sheet]
                matched, unmatched = run_match(ws1, perf_records, year, month)
                seg_results[sheet] = {'matched': matched, 'unmatched': unmatched}

        st.session_state['match_cache_key'] = cache_key
        st.session_state['seg_results']     = seg_results
        st.session_state['perf_cnt']        = len(perf_records)
        st.session_state['year']            = year
        st.session_state['month']           = month

    seg_results = st.session_state['seg_results']
    perf_cnt    = st.session_state['perf_cnt']
    year        = st.session_state['year']
    month       = st.session_state['month']

    total_matched   = sum(len(v['matched'])   for v in seg_results.values())
    total_unmatched = sum(len(v['unmatched']) for v in seg_results.values())

    # ── 요약 지표 ─────────────────────────────────────────────────────────────
    st.markdown("---")
    mc1, mc2, mc3, mc4 = st.columns(4)
    with mc1:
        st.markdown(f"""<div class="metric-card">
            <div class="mc-label">대상 연월</div>
            <div class="mc-value sm">{year}년 {month}월</div>
        </div>""", unsafe_allow_html=True)
    with mc2:
        st.markdown(f"""<div class="metric-card">
            <div class="mc-label">분석 구간</div>
            <div class="mc-value sm">{' · '.join(seg_results.keys())}</div>
        </div>""", unsafe_allow_html=True)
    with mc3:
        cls = 'green' if total_matched else 'amber'
        st.markdown(f"""<div class="metric-card">
            <div class="mc-label">총 매칭 건수</div>
            <div class="mc-value {cls}">{total_matched}</div>
        </div>""", unsafe_allow_html=True)
    with mc4:
        cls = 'amber' if total_unmatched else ''
        st.markdown(f"""<div class="metric-card">
            <div class="mc-label">미매칭 건수</div>
            <div class="mc-value {cls}">{total_unmatched}</div>
        </div>""", unsafe_allow_html=True)

    # ── 구간별 결과 탭 ────────────────────────────────────────────────────────
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    tab_labels = list(seg_results.keys())
    tabs = st.tabs([f"📍 {t}" for t in tab_labels])

    for tab, sheet in zip(tabs, tab_labels):
        with tab:
            data    = seg_results[sheet]
            matched = data['matched']
            unmatched = data['unmatched']
            matched_df   = pd.DataFrame(matched)
            unmatched_df = pd.DataFrame(unmatched)

            m_cnt = len(matched)
            u_cnt = len(unmatched)
            m_cls = 'has-match' if m_cnt else ''
            st.markdown(f"""
            <div style="display:flex; gap:16px; align-items:center; margin:12px 0 8px;">
                <span style="font-size:0.78rem; font-weight:700; color:#0f766e;">
                    ✅ 매칭 {m_cnt}건
                </span>
                <span style="font-size:0.78rem; color:{'#b45309' if u_cnt else '#94a3b8'};">
                    {"⚠️" if u_cnt else "—"} 미매칭 {u_cnt}건
                </span>
            </div>
            """, unsafe_allow_html=True)

            # 매칭 결과
            if matched_df.empty:
                st.markdown('<div class="warn-box">매칭된 항목이 없습니다. 날짜와 구간(km 기준)이 겹치는 경우에만 매칭됩니다.</div>', unsafe_allow_html=True)
            else:
                st.dataframe(
                    matched_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config=COLUMN_CONFIG,
                )

            # 미매칭
            if not unmatched_df.empty:
                with st.expander(f"⚠️ 미매칭 레일연마 {u_cnt}건 — 일정엔 있으나 실적 구간과 겹치지 않음"):
                    st.dataframe(unmatched_df, use_container_width=True, hide_index=True)


# ── 라우터 ────────────────────────────────────────────────────────────────────

if st.session_state['page'] == 'home':
    page_home()
elif st.session_state['page'] == 'match':
    page_match()
