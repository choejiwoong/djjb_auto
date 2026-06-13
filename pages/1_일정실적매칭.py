"""
일정-실적 매칭 페이지
"""

import io

import openpyxl
import pandas as pd
import streamlit as st

from utils.auth import check_password
from utils.match_logic import extract_year_month, load_perf, run_match
from utils.styles import COMMON_CSS

# ── 페이지 설정 ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="일정-실적 매칭 | 레일연마차 업무자동화",
    page_icon="🔍",
    layout="wide",
)

# ── 로그인 체크 (pages에서도 반드시 체크) ────────────────────────────────────

if not check_password():
    st.stop()

st.markdown(COMMON_CSS, unsafe_allow_html=True)

# ── 사이드바 ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div class="sb-brand">
        <div class="sb-icon">🚃</div>
        <div class="sb-title">레일연마차<br>업무자동화</div>
        <div class="sb-sub">대저궤도장비분소</div>
    </div>
    """, unsafe_allow_html=True)

# ── 컬럼 설정 ─────────────────────────────────────────────────────────────────

COLUMN_CONFIG = {
    '일자':      st.column_config.TextColumn('일자', width=130),
    '연마차명':  st.column_config.TextColumn('연마차명', width=150),
    '공종':      st.column_config.TextColumn('공종', width=120),
    '분소':      st.column_config.TextColumn('분소', width=80),
    '일정_시점': st.column_config.TextColumn('일정 시점', width=100),
    '일정_종점': st.column_config.TextColumn('일정 종점', width=100),
    '실적_구간': st.column_config.TextColumn('실적 구간', width=140),
    '실적_부터': st.column_config.TextColumn('실적 부터', width=100),
    '실적_까지': st.column_config.TextColumn('실적 까지', width=100),
}

# ── 페이지 헤더 ───────────────────────────────────────────────────────────────

st.markdown("""
<div class="page-header">
    <div class="ph-badge">일정-실적 매칭</div>
    <h2>🔍 일정-실적 구간 매칭</h2>
    <p>두 파일을 업로드하면 1구간·2구간을 자동으로 매칭합니다.</p>
</div>
""", unsafe_allow_html=True)

# ── 파일 업로드 ───────────────────────────────────────────────────────────────

col_up1, col_up2 = st.columns(2)

with col_up1:
    st.markdown('<div class="upload-section"><div class="us-title">📋 파일 업로드</div>', unsafe_allow_html=True)
    f_schedule = st.file_uploader(
        "궤도정비공사일정 파일",
        type=['xlsx'], key="schedule",
        help="0000년 0월 궤도시설 보수정비공사 일정.xlsx",
    )
    f_perf = st.file_uploader(
        "레일연마차 작업 및 검수실적 파일",
        type=['xlsx'], key="perf",
        help="1. 0000년 0월 레일연마차 작업 및 검수실적.xlsx",
    )
    st.markdown('</div>', unsafe_allow_html=True)

with col_up2:
    st.markdown('<div class="upload-section"><div class="us-title">📂 업로드 상태</div>', unsafe_allow_html=True)

    if f_schedule:
        st.markdown(f'<div class="status-row ok"><div class="dot green"></div><b>일정 파일</b> — {f_schedule.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-row wait"><div class="dot gray"></div>궤도정비공사일정 파일 대기 중</div>', unsafe_allow_html=True)

    if f_perf:
        st.markdown(f'<div class="status-row ok"><div class="dot green"></div><b>실적 파일</b> — {f_perf.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-row wait"><div class="dot gray"></div>레일연마차 작업 및 검수실적 파일 대기 중</div>', unsafe_allow_html=True)

    if f_schedule and f_perf:
        year, month = extract_year_month(f_schedule.name)
        st.markdown(f"""
        <div style="margin-top:12px; padding:10px 14px; background:#f0fdf4;
                    border-radius:8px; border:1px solid #bbf7d0;
                    font-size:0.82rem; color:#0f766e;">
            ✅ 두 파일이 준비됐습니다. <b>{year}년 {month}월</b> 데이터를 분석합니다.<br>
            <span style="font-size:0.75rem; color:#16a34a; margin-top:4px; display:block;">
            1구간 · 2구간 자동 매칭 실행 중...
            </span>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="margin-top:12px; padding:10px 14px; background:#f8fafc;
                    border-radius:8px; border:1px solid #e2e8f0;
                    font-size:0.82rem; color:#94a3b8;">
            두 파일을 모두 업로드하면 자동으로 매칭이 시작됩니다.
        </div>
        """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ── 파일 없으면 중단 ─────────────────────────────────────────────────────────

if not f_schedule or not f_perf:
    st.stop()

# ── 매칭 실행 (파일명이 바뀔 때만 재실행) ────────────────────────────────────

cache_key = f"{f_schedule.name}_{f_perf.name}"
if st.session_state.get('match_cache_key') != cache_key:
    try:
        wb1 = openpyxl.load_workbook(io.BytesIO(f_schedule.read()), data_only=True)
        wb2 = openpyxl.load_workbook(io.BytesIO(f_perf.read()),     data_only=True)
    except Exception as e:
        st.error(f"파일을 열 수 없습니다: {e}")
        st.stop()

    year, month = extract_year_month(f_schedule.name)
    _, _, perf_records = load_perf(wb2['장비별 세부작업실적'])

    seg_results = {}
    for sheet in ['1구간', '2구간']:
        if sheet in wb1.sheetnames:
            matched, unmatched = run_match(wb1[sheet], perf_records, year, month)
            seg_results[sheet] = {'matched': matched, 'unmatched': unmatched}

    st.session_state['match_cache_key'] = cache_key
    st.session_state['seg_results']     = seg_results
    st.session_state['perf_cnt']        = len(perf_records)
    st.session_state['year']            = year
    st.session_state['month']           = month

seg_results = st.session_state['seg_results']
year        = st.session_state['year']
month       = st.session_state['month']
perf_cnt    = st.session_state['perf_cnt']

total_matched   = sum(len(v['matched'])   for v in seg_results.values())
total_unmatched = sum(len(v['unmatched']) for v in seg_results.values())

# ── 요약 지표 ─────────────────────────────────────────────────────────────────

st.markdown("---")
mc1, mc2, mc3, mc4 = st.columns(4)
with mc1:
    st.markdown(f"""<div class="metric-card">
        <div class="mc-label">대상 연월</div>
        <div class="mc-value sm">{year}년 {month}월</div>
    </div>""", unsafe_allow_html=True)
with mc2:
    st.markdown(f"""<div class="metric-card">
        <div class="mc-label">분석 구간</div>
        <div class="mc-value sm">{' · '.join(seg_results.keys())}</div>
    </div>""", unsafe_allow_html=True)
with mc3:
    cls = 'green' if total_matched else 'amber'
    st.markdown(f"""<div class="metric-card">
        <div class="mc-label">총 매칭 건수</div>
        <div class="mc-value {cls}">{total_matched}</div>
    </div>""", unsafe_allow_html=True)
with mc4:
    cls = 'amber' if total_unmatched else ''
    st.markdown(f"""<div class="metric-card">
        <div class="mc-label">미매칭 건수</div>
        <div class="mc-value {cls}">{total_unmatched}</div>
    </div>""", unsafe_allow_html=True)

# ── 구간별 결과 탭 ────────────────────────────────────────────────────────────

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
tabs = st.tabs([f"📍 {sheet}" for sheet in seg_results])

for tab, sheet in zip(tabs, seg_results):
    with tab:
        matched   = seg_results[sheet]['matched']
        unmatched = seg_results[sheet]['unmatched']
        matched_df   = pd.DataFrame(matched)
        unmatched_df = pd.DataFrame(unmatched)
        m_cnt, u_cnt = len(matched), len(unmatched)

        st.markdown(f"""
        <div style="display:flex; gap:16px; align-items:center; margin:12px 0 8px;">
            <span style="font-size:0.78rem; font-weight:700; color:#0f766e;">
                ✅ 매칭 {m_cnt}건
            </span>
            <span style="font-size:0.78rem; color:{'#b45309' if u_cnt else '#94a3b8'};">
                {"⚠️" if u_cnt else "—"} 미매칭 {u_cnt}건
            </span>
        </div>
        """, unsafe_allow_html=True)

        if matched_df.empty:
            st.markdown('<div class="warn-box">매칭된 항목이 없습니다. 날짜와 구간(km 기준)이 겹치는 경우에만 매칭됩니다.</div>', unsafe_allow_html=True)
        else:
            st.dataframe(matched_df, use_container_width=True, hide_index=True, column_config=COLUMN_CONFIG)

        if not unmatched_df.empty:
            with st.expander(f"⚠️ 미매칭 레일연마 {u_cnt}건 — 일정엔 있으나 실적 구간과 겹치지 않음"):
                st.dataframe(unmatched_df, use_container_width=True, hide_index=True)
